"""Sparse, source-neutral workbook regions for memo composition."""
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class WorkbookRegion:
    category: str
    sheet: str
    cell_range: str
    headers: tuple[object, ...]
    rows: tuple[tuple[object, ...], ...]
    formula_count: int
    lineage: tuple[str, ...]
    source_uri: str


@dataclass(frozen=True)
class WorkbookProfile:
    source_uri: str
    regions: tuple[WorkbookRegion, ...]


def profile_workbook(path, max_rows=40, max_columns=16):
    source = Path(path).resolve()
    formulas = load_workbook(source, data_only=False, read_only=False, keep_links=True)
    values = load_workbook(source, data_only=True, read_only=False, keep_links=True)
    regions = []
    try:
        for sheet in formulas.worksheets:
            region = _profile_sheet(sheet, values[sheet.title], max_rows, max_columns, source)
            if region:
                regions.append(region)
    finally:
        formulas.close()
        values.close()
    return WorkbookProfile(str(source), tuple(regions))


def _profile_sheet(sheet, value_sheet, max_rows, max_columns, source):
    cells = [cell for cell in sheet._cells.values() if cell.value not in (None, "")]
    if not cells:
        return None
    columns = _bounded_axis(Counter(cell.column for cell in cells), max_columns)
    rows = _bounded_rows(cells, columns, max_rows)
    if not rows or not columns:
        return None
    table_rows = tuple(
        tuple(_cell_display(sheet, value_sheet, row, column) for column in columns)
        for row in rows
    )
    first, last = f"{get_column_letter(columns[0])}{rows[0]}", f"{get_column_letter(columns[-1])}{rows[-1]}"
    lineage = tuple(
        f"{sheet.title}!{get_column_letter(column)}{row}"
        for row in rows for column in columns
        if (cell := sheet._cells.get((row, column))) is not None and cell.value not in (None, "")
    )
    context = " ".join(str(value) for value in table_rows[0] if value not in (None, ""))
    return WorkbookRegion(
        category=_category(sheet.title, context),
        sheet=sheet.title,
        cell_range=f"{first}:{last}",
        headers=table_rows[0],
        rows=table_rows,
        formula_count=sum(cell.data_type == "f" or _is_formula(cell.value) for cell in cells),
        lineage=lineage,
        source_uri=str(source),
    )


def _bounded_axis(counts, limit):
    dense = sorted(axis for axis, count in counts.items() if count >= 2)
    if dense:
        candidates = [axis for axis in counts if dense[0] <= axis <= dense[-1]]
    else:
        candidates = list(counts)
    chosen = sorted(candidates, key=lambda axis: (-counts[axis], axis))[:limit]
    return tuple(sorted(chosen))


def _bounded_rows(cells, columns, limit):
    selected = set(columns)
    counts = Counter(cell.row for cell in cells if cell.column in selected)
    meaningful = [row for row, count in sorted(counts.items()) if count >= 2]
    return tuple((meaningful or sorted(counts))[:limit])


def _cell_display(sheet, value_sheet, row, column):
    cell = sheet._cells.get((row, column))
    if cell is None or cell.value in (None, ""):
        return None
    value = cell.value
    if cell.data_type == "f" or _is_formula(value):
        cached = value_sheet._cells.get((row, column))
        value = cached.value if cached is not None and not _is_formula(cached.value) else None
        return "待重算" if value is None else _scalar(value)
    return _scalar(value)


def _is_formula(value):
    if isinstance(value, str):
        return value.startswith("=")
    kind = str(getattr(value, "t", "")).casefold()
    return "formula" in kind


def _scalar(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value if isinstance(value, (str, int, float, bool)) or value is None else str(value)


def _category(sheet_name, header_text):
    text = f"{sheet_name} {header_text}".casefold()
    if any(word in text for word in ("sensitive", "sensitivity", "敏感性")):
        return "sensitivity"
    if any(word in text for word in ("return", "irr", "moic", "收益", "回报")):
        return "returns"
    if any(word in text for word in ("transaction", "term", "deal", "sources & uses", "交易", "条款", "投资方案")):
        return "transaction"
    if any(word in text for word in ("cap table", "ownership", "shareholding", "股权结构", "持股")):
        return "ownership"
    if any(word in text for word in ("comparable", "trading comps", "可比公司")):
        return "comparables"
    if (re.search(r"20\d{2}\s*e\b", text)
            or any(word in text for word in ("forecast", "p&l", "income statement", "balance sheet",
                                              "cashflow", "cash flow", "opex", "product buildup", "预测"))):
        return "financial_forecast"
    if any(word in text for word in ("historical", "actual", "财务", "损益", "资产负债")):
        return "financial_history"
    if any(word in text for word in ("summary", "overview", "dashboard", "概要", "概览")):
        return "overview"
    return "evidence"
