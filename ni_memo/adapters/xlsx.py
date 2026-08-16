"""Strict, coordinate-based XLSX fact extraction."""
import re
from pathlib import Path

from openpyxl import load_workbook

from ni_memo.facts import FactRecord, FactStatus


def extract_mapped_facts(workbook_path, mapping):
    path = Path(workbook_path)
    wb = load_workbook(path, read_only=True, data_only=False)
    wb_values = load_workbook(path, read_only=True, data_only=True)
    facts = []
    try:
        for rule in mapping:
            sheet = rule.get("sheet", "")
            cell = rule.get("cell", "")
            if sheet not in wb.sheetnames:
                raise ValueError(f"sheet not found: {sheet}")
            if not cell:
                raise ValueError(f"cell is required for {rule.get('key', '')}")
            raw = wb[sheet][cell].value
            if raw is None:
                facts.append(_fact(path, rule, None, FactStatus.MISSING, "cell"))
                continue
            value = raw
            method = "formula" if isinstance(raw, str) and raw.startswith("=") else "cell"
            if method == "formula":
                cached = wb_values[sheet][cell].value
                if cached is not None:
                    value = cached
                    method = "formula_cached"
            pattern = rule.get("extract_regex")
            if pattern:
                matches = re.findall(pattern, str(raw))
                if len(matches) != 1:
                    raise ValueError(f"extract_regex must match exactly once for {rule['key']}")
                value = matches[0][0] if isinstance(matches[0], tuple) else matches[0]
                value = float(value) if "." in str(value) else int(value)
                method = "cell_regex"
            if rule.get("scale") is not None:
                if not isinstance(value, (int, float)) or isinstance(value, bool):
                    raise ValueError(f"scale requires a numeric value for {rule['key']}")
                value = value * float(rule["scale"])
                method += "_scaled"
            facts.append(_fact(path, rule, value, FactStatus.SOURCE_ONLY, method, raw))
    finally:
        wb.close()
        wb_values.close()
    return facts


def _fact(path, rule, value, status, method, raw=None):
    location = f"{rule.get('sheet', '')}!{rule.get('cell', '')}"
    evidence = () if raw is None else (f"raw={raw}",)
    return FactRecord(
        key=rule["key"], value=value, unit=rule.get("unit"), as_of=rule.get("as_of"),
        source_type="xlsx", source_uri=str(path), source_location=location,
        extraction_method=method, confidence=rule.get("confidence", "high"),
        status=status, evidence=evidence,
    )
